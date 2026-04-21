from pathlib import Path

from openpyxl import load_workbook

from src.ingestion.models import ParsedDocument, ParsedPage
from src.ingestion.parsers.base import DocumentParser


class XLSXParser(DocumentParser):
    """
    Extracts text from Excel (.xlsx) files using openpyxl.

    Each worksheet becomes one "page". Rows are converted to
    pipe-delimited text lines, preserving the tabular structure
    that makes financial data meaningful.

    Why pipe-delimited?
    Commas appear inside financial numbers ("1,234,567") which breaks
    CSV parsing. Pipes are rarely used in financial data.

    Example output for a P&L sheet:
        Revenue | Q1 | Q2 | Q3 | Q4
        Product A | 1234 | 1456 | 1678 | 1890
        Product B | 567 | 589 | 612 | 634
        Total | 1801 | 2045 | 2290 | 2524

    Limitations:
    - Formulas are not evaluated — values are read (openpyxl data_only=True)
    - Merged cells only return value in the top-left cell
    - Charts are ignored (no text layer)
    - Hidden rows/columns are included by default
    """

    def parse(self, file_path: Path) -> ParsedDocument:
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"XLSX not found: {file_path}")

        file_hash = self.compute_file_hash(file_path)

        # data_only=True reads computed cell values, not formulas
        wb = load_workbook(str(file_path), data_only=True)

        pages: list[ParsedPage] = []

        for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            lines = []

            for row in ws.iter_rows(values_only=True):
                # Skip entirely empty rows
                if all(cell is None for cell in row):
                    continue

                # Convert each cell to string, replace None with empty string
                cells = [str(cell) if cell is not None else "" for cell in row]
                lines.append(" | ".join(cells))

            content = "\n".join(lines).strip()
            if content:
                pages.append(ParsedPage(
                    content=content,
                    page_number=sheet_num,
                    section_title=sheet_name,   # sheet name as section
                    metadata={"sheet_name": sheet_name},
                ))

        if not pages:
            raise ValueError(f"No data found in {file_path.name}")

        return ParsedDocument(
            filename=file_path.name,
            file_path=str(file_path),
            doc_type="xlsx",
            pages=pages,
            file_hash=file_hash,
        )
